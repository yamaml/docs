#!/usr/bin/env python3
"""Generate downloadable SimpleDSP example files in TSV, CSV, and styled XLSX.

All three examples shown in the SimpleDSP specification are produced here from
a single canonical data source, ensuring the downloadable files match the
inline documentation exactly.

The XLSX files are styled for visual comfort:
  - Block headers ([@NS], [MAIN], …) as merged dark-slate bands with white text
  - Comment header rows (#Name Property …) in light grey with italic text
  - ID rows highlighted in warm amber
  - Data rows alternating between white and very light grey
  - Consistent thin borders, proper column widths

The generated XLSX is post-processed to strip t='n' from valueless cells for
strict Excel file-validation compatibility.
"""
import csv
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / 'static' / 'examples' / 'simpledsp'
NCOLS = 7

# ── Colors ───────────────────────────────────────────────────────
BLOCK_BG, BLOCK_FG = '2E4A62', 'FFFFFF'
COMMENT_BG, COMMENT_FG = 'E8ECEF', '6B7280'
ID_BG, ID_FG = 'FEF3C7', '78350F'
DATA_BG_A, DATA_BG_B = 'FFFFFF', 'F9FAFB'
DATA_FG = '1F2937'
BORDER_COLOR = 'D1D5DB'

# ── Styles ───────────────────────────────────────────────────────
font_block = Font(name='Calibri', size=12, bold=True, color=BLOCK_FG)
font_comment = Font(name='Consolas', size=10, italic=True, color=COMMENT_FG)
font_id = Font(name='Calibri', size=11, bold=True, color=ID_FG)
font_data = Font(name='Calibri', size=11, color=DATA_FG)

fill_block = PatternFill('solid', fgColor=BLOCK_BG)
fill_comment = PatternFill('solid', fgColor=COMMENT_BG)
fill_id = PatternFill('solid', fgColor=ID_BG)
fill_data_a = PatternFill('solid', fgColor=DATA_BG_A)
fill_data_b = PatternFill('solid', fgColor=DATA_BG_B)

thin = Side(border_style='thin', color=BORDER_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')

DEFAULT_WIDTHS = {'A': 20, 'B': 24, 'C': 6, 'D': 6, 'E': 12, 'F': 18, 'G': 40}


def classify(row):
    if not row:
        return 'blank'
    first = (row[0] or '').strip()
    if first.startswith('['):
        return 'block'
    if first.startswith('#'):
        return 'comment'
    # ID row: either "ID" as the name, or name ending with "ID" with value type "ID" in column 5
    if len(row) >= 5 and row[4] == 'ID':
        return 'id'
    return 'data'


def fmt(cell, kind, c_idx, parity):
    cell.border = border
    if kind == 'block':
        cell.fill, cell.font, cell.alignment = fill_block, font_block, align_left
    elif kind == 'comment':
        cell.fill, cell.font, cell.alignment = fill_comment, font_comment, align_left
    elif kind == 'id':
        cell.fill, cell.font = fill_id, font_id
        cell.alignment = align_center if c_idx in (3, 4) else align_left
    else:
        cell.fill = fill_data_a if parity == 0 else fill_data_b
        cell.font = font_data
        cell.alignment = align_center if c_idx in (3, 4) else align_left


def write_tsv(rows, path):
    with path.open('w', encoding='utf-8', newline='') as f:
        prev = None
        for r in rows:
            kind = classify(r)
            if kind == 'block' and prev is not None:
                f.write('\n')
            f.write('\t'.join(r) + '\n')
            prev = kind


def write_csv(rows, path):
    with path.open('w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        prev = None
        for r in rows:
            kind = classify(r)
            if kind == 'block' and prev is not None:
                w.writerow([])
            w.writerow(r)
            prev = kind


def write_xlsx(rows, path, widths=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'SimpleDSP'

    merges = []
    parity = 0
    out_row = 0
    prev_block_idx = None  # 1-based output row of the last [block] header

    for src_row in rows:
        kind = classify(src_row)
        if kind == 'blank':
            continue
        out_row += 1

        if kind == 'block':
            cell = ws.cell(row=out_row, column=1, value=src_row[0])
            fmt(cell, kind, 1, 0)
            for c in range(2, NCOLS + 1):
                fmt(ws.cell(row=out_row, column=c, value=' '), kind, c, 0)
            merges.append(f'A{out_row}:{get_column_letter(NCOLS)}{out_row}')
            prev_block_idx = out_row
            parity = 0
            continue

        # Normal row — write all 7 cells
        for c_idx in range(1, NCOLS + 1):
            raw = src_row[c_idx - 1] if c_idx - 1 < len(src_row) else ''
            value = raw if raw != '' else ' '
            cell = ws.cell(row=out_row, column=c_idx, value=value)
            fmt(cell, kind, c_idx, parity % 2)
        if kind == 'data':
            parity += 1

    for mr in merges:
        ws.merge_cells(mr)

    for col, w in (widths or DEFAULT_WIDTHS).items():
        ws.column_dimensions[col].width = w

    wb.save(path)

    # Post-process: strip t="n" from valueless cells
    tmp = path.with_suffix('.tmp.xlsx')
    with zipfile.ZipFile(path, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item == 'xl/worksheets/sheet1.xml':
                    xml = data.decode('utf-8')
                    xml = re.sub(r'(<c\s[^>]*?)\s+t="n"(\s*/>|\s*></c>)', r'\1\2', xml)
                    data = xml.encode('utf-8')
                zout.writestr(item, data)
    shutil.move(tmp, path)


def generate(name, rows, widths=None):
    tsv = OUT_DIR / f'{name}.tsv'
    csv_ = OUT_DIR / f'{name}.csv'
    xlsx = OUT_DIR / f'{name}.xlsx'
    write_tsv(rows, tsv)
    write_csv(rows, csv_)
    write_xlsx(rows, xlsx, widths)
    print(f"  {name}: tsv={tsv.stat().st_size}B csv={csv_.stat().st_size}B xlsx={xlsx.stat().st_size}B")


# ──────────────────────────────────────────────────────────────────
# Example 1: The Big Bang Theory characters (Section 8 demo)
# ──────────────────────────────────────────────────────────────────
tbbt_rows = [
    ['[@NS]'],
    ['schema', 'http://schema.org/'],
    ['@base', 'http://purl.org/yama/examples/2022/tbbt/0.1/'],
    ['[MAIN]'],
    ['#Name', 'Property', 'Min', 'Max', 'ValueType', 'Constraint', 'Comment'],
    ['ID', 'foaf:Person', '1', '1', 'ID', '', 'Unique identifier for the character'],
    ['Name', 'foaf:name', '1', '1', 'literal', 'xsd:string', 'Full name of the character'],
    ['Family Name', 'foaf:familyName', '1', '1', 'literal', 'xsd:string', 'Family name'],
    ['First Name', 'foaf:firstName', '1', '1', 'literal', 'xsd:string', 'Given name'],
    ['Job Title', 'schema:jobTitle', '0', '1', 'literal', 'xsd:string', 'Job title of the character'],
    ['Parents', 'schema:parent', '0', '-', 'IRI', '', 'Parents of the character'],
    ['Children', 'schema:children', '0', '-', 'IRI', '', 'Children of the character'],
    ['Knows', 'foaf:knows', '0', '-', 'IRI', '', 'Other characters this character knows'],
    ['Wikidata', 'rdfs:seeAlso', '0', '-', 'IRI', '', 'Wikidata entity for this character'],
    ['Home Address', 'schema:address', '0', '1', 'structured', '#address', 'Home address of the character'],
    ['Portrayed by', 'schema:byArtist', '0', '1', 'IRI', '', 'Actor who portrays this character'],
    ['[address]'],
    ['#Name', 'Property', 'Min', 'Max', 'ValueType', 'Constraint', 'Comment'],
    ['Street', 'schema:streetAddress', '0', '1', 'literal', 'xsd:string', 'Building and street address'],
    ['Locality', 'schema:addressLocality', '0', '1', 'literal', 'xsd:string', 'City or town'],
    ['Region', 'schema:addressRegion', '0', '1', 'literal', 'xsd:string', 'State, province, or region'],
    ['Country', 'schema:addressCountry', '0', '1', 'literal', 'xsd:string', 'Country'],
    ['Postal Code', 'schema:postalCode', '0', '1', 'literal', 'xsd:string', 'Postal or ZIP code'],
]

# ──────────────────────────────────────────────────────────────────
# Example 2: Manga Design Metadata (DMM project)
# ──────────────────────────────────────────────────────────────────
manga_rows = [
    ['[@NS]'],
    ['foaf', 'http://xmlns.com/foaf/0.1/'],
    ['dc', 'http://purl.org/dc/elements/1.1/'],
    ['dct', 'http://purl.org/dc/terms/'],
    ['dmm', 'http://anise.slis.tsukuba.ac.jp/dmm/'],
    ['rdfs', 'http://www.w3.org/2000/01/rdf-schema#'],
    ['[MAIN]'],
    ['#項目規則名', 'プロパティ', '最少', '最大', '値タイプ', '値制約', 'コメント'],
    ['MangaID', 'dmm:Manga', '1', '1', 'ID'],
    ['Setting', 'dmm:hasSetting', '0', '-', '構造化', '#Setting', 'マンガの設定'],
    ['[Setting]'],
    ['SettingID', 'dmm:Setting', '1', '1', 'ID'],
    ['title', 'dc:title', '0', '1', '文字列'],
    ['image', 'dmm:image', '0', '-', '参照値'],
    ['description', 'dc:description', '0', '1', '文字列'],
    ['relation', 'dmm:relation', '0', '-', '構造化', '#Setting'],
    ['CharacterSetting', 'dmm:hasCharacterSetting', '0', '1', '構造化', '#CharacterSetting', '人物設定'],
    ['StageSetting', 'dmm:hasStageSetting', '0', '1', '構造化', '#StageSetting', '舞台設定'],
    ['PlotSetting', 'dmm:hasPlotSetting', '0', '1', '構造化', '#PlotSetting', '事柄設定'],
    ['[CharacterSetting]'],
    ['CharacterSettingID', 'dmm:CharacterSetting', '1', '1', 'ID'],
    ['name', 'dmm:name', '0', '-', '文字列'],
    ['sex', 'dmm:sex', '0', '-', '文字列'],
    ['hometown', 'dmm:hometown', '0', '-', '文字列'],
    ['model', 'dmm:model', '0', '-', '参照値'],
    ['role', 'dmm:hasRole', '0', '-', '文字列'],
    ['[PlotSetting]'],
    ['PlotSettingID', 'dmm:StorySetting', '1', '1', 'ID'],
    ['cast', 'dmm:hasCarahcter', '0', '-', '構造化', '#CharacterSetting'],
    ['stage', 'dmm:hasStage', '0', '-', '構造化', '#StageSetting'],
    ['[StageSetting]'],
    ['StageSettingID', 'dmm:StageSetting', '1', '1', 'ID'],
    ['location', 'dmm:name', '0', '-', '文字列'],
    ['coord', 'dmm:coord', '0', '-', '文字列'],
    ['[Role]'],
    ['RoleID', 'dmm:Role', '1', '1', 'ID'],
    ['RoleName', 'dmm:name', '0', '-', '文字列'],
    ['description', 'dc:description', '0', '-', '文字列'],
]

# ──────────────────────────────────────────────────────────────────
# Example 3: NDL bibliographic record (canonical example from §6.2.5)
# ──────────────────────────────────────────────────────────────────
ndl_rows = [
    ['[@NS]'],
    ['dcndl', 'http://ndl.go.jp/dcndl/terms/'],
    ['ndlsh', 'http://id.ndl.go.jp/auth/ndlsh/'],
    ['bsh', 'http://id.ndl.go.jp/auth/bsh/'],
    ['ndlbooks', 'http://iss.ndl.go.jp/books/'],
    ['@base', 'http://ndl.go.jp/dcndl/dsp/biblio'],
    ['[MAIN]'],
    ['#項目規則名', 'プロパティ', '最小', '最大', '値タイプ', '値制約', '説明'],
    ['書誌ID', 'foaf:Document', '1', '1', 'ID', 'ndlbooks:', '文書のID'],
    ['タイトル', 'dcterms:title', '1', '1', '構造化', '#構造化タイトル', '文書の表題'],
    ['著者', 'dcterms:creator', '0', '1', '構造化', 'foaf:Agent', '文書の作者'],
    ['発行日', 'dcterms:issued', '1', '1', '文字列', 'xsd:date', '文書の発行日'],
    ['主題', 'dcterms:subject', '0', '-', '参照値', 'ndlsh: bsh:', '文書の主題'],
    ['[構造化タイトル]'],
    ['#項目規則名', 'プロパティ', '最小', '最大', '値タイプ', '値制約', '説明'],
    ['リテラル値', 'xl:literalForm', '1', '1', '文字列', '', 'タイトル自身'],
    ['読み', 'dcndl:transcription', '0', '1', '文字列', '', 'タイトルの読み'],
]


if __name__ == '__main__':
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Generating SimpleDSP example files…")
    generate('tbbt-characters', tbbt_rows)
    generate('manga-design', manga_rows, widths={'A': 22, 'B': 26, 'C': 6, 'D': 6, 'E': 10, 'F': 22, 'G': 18})
    generate('ndl-bibliographic', ndl_rows, widths={'A': 16, 'B': 22, 'C': 6, 'D': 6, 'E': 10, 'F': 20, 'G': 18})
    print("Done.")
